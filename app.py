# --- v20.2 (Base v19.12 + Correzione Indentazione Debug + Aggiunta Percorso Critico) ---
import streamlit as st
from lxml import etree
import pandas as pd
from datetime import datetime, date, timedelta
import re
import isodate
from io import BytesIO
import math
import plotly.graph_objects as go
import traceback
import os
import locale
try:
    import plotly.io as pio
    from openpyxl.drawing.image import Image
    _kaleido_installed = True
except ImportError:
    _kaleido_installed = False
import openpyxl.utils
import plotly.express as px

# --- Imposta Locale Italiano ---
# ... (Codice invariato v17.13) ...
_locale_warning_shown = False
try: locale.setlocale(locale.LC_TIME, 'it_IT.UTF-8')
except locale.Error:
    try: locale.setlocale(locale.LC_TIME, 'italian')
    except locale.Error:
        try: locale.setlocale(locale.LC_TIME, '')
        except locale.Error:
             if not _locale_warning_shown:
                print("WARNING: Impossibile impostare qualsiasi locale per i nomi dei mesi.")
                _locale_warning_shown = True

# --- CONFIGURAZIONE DELLA PAGINA ---
st.set_page_config(page_title="InfraTrack v20.2", page_icon="🚆", layout="wide") # Version updated

# --- CSS ---
# ... (CSS invariato v17.12) ...
st.markdown("""
<style>
     .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6, .stApp p, .stApp .stDataFrame, .stApp .stButton>button { font-size: 0.85rem !important; }
    .stApp h2 { font-size: 1.5rem !important; }
    .stApp .stMarkdown h4 { font-size: 1.1rem !important; margin-bottom: 0.5rem; margin-top: 1rem; }
    .stApp .stMarkdown h5 { font-size: 0.90rem !important; margin-bottom: 0.5rem; margin-top: 0.8rem; }
    .stApp .stMarkdown h6 { font-size: 0.88rem !important; margin-bottom: 0.4rem; margin-top: 0.8rem; font-weight: bold;}
    button[data-testid="stButton"][kind="primary"][key="reset_button"],
    button[data-testid="stButton"][kind="secondary"][key="clear_cache_button"] { padding: 0.2rem 0.5rem !important; line-height: 1.2 !important; font-size: 1.0rem !important; border-radius: 0.25rem !important; margin-right: 5px; }
    button[data-testid="stButton"][kind="primary"][key="reset_button"]:disabled { cursor: not-allowed; opacity: 0.5; }
    .stApp { padding-top: 2rem; }
    .stDataFrame td { text-align: center !important; }
    .stDataFrame th:nth-child(4), .stDataFrame td:nth-child(4) { text-align: left !important; } /* Colonna Riepilogo SIL */
    .stDataFrame th:nth-child(2), .stDataFrame td:nth-child(2) { text-align: center !important; } /* Durata TUP/TUF */
    .stDataFrame th:nth-child(3), .stDataFrame td:nth-child(3) { text-align: left !important; } /* Nome Mezzo */
    div[data-testid="stDateInput"] label { font-size: 0.85rem !important; }
    div[data-testid="stDateInput"] input { font-size: 0.85rem !important; padding: 0.3rem 0.5rem !important;}
    .stCaptionContainer { font-size: 0.75rem !important; margin-top: -0.5rem; margin-bottom: 1rem;}
    .progress-text { font-size: 0.8rem; color: grey; margin-left: 10px; }
    /* Tabella Pivot Mezzi: allinea nomi mezzi a sinistra */
    .stDataFrame thead th:not(:first-child) { text-align: center !important; } /* Intestazioni colonne pivot */
    .stDataFrame tbody th { text-align: left !important; } /* Indice righe pivot (Periodo) */
</style>
""", unsafe_allow_html=True)


# --- TITOLO E HEADER ---
st.markdown("## 🚆 InfraTrack v20.2") # Version updated
st.caption("La tua centrale di controllo per progetti infrastrutturali")

# --- GESTIONE RESET E CACHE ---
# ... (Codice invariato v17.9) ...
if 'widget_key_counter' not in st.session_state: st.session_state.widget_key_counter = 0
if 'file_processed_success' not in st.session_state: st.session_state.file_processed_success = False
col_btn_1, col_btn_2, col_btn_3 = st.columns([0.1, 0.2, 0.7])
with col_btn_1:
    if st.button("🔄", key="reset_button", help="Resetta l'analisi (Svuota Sessione e File)", disabled=not st.session_state.file_processed_success):
        st.session_state.widget_key_counter += 1; st.session_state.file_processed_success = False
        if 'uploaded_file_state' in st.session_state: del st.session_state['uploaded_file_state']
        keys_to_reset = list(st.session_state.keys())
        for key in keys_to_reset:
            if not key.startswith("_") and key != 'widget_key_counter': del st.session_state[key]
        st.toast("Sessione resettata.", icon="🔄"); st.rerun()
with col_btn_2:
    if st.button("🗑️ Svuota Cache", key="clear_cache_button", help="Elimina i dati temporanei calcolati (Forza ri-analisi @st.cache_data)"):
        st.cache_data.clear(); st.toast("Cache dei dati svuotata! I dati verranno ricalcolati alla prossima analisi.", icon="✅")

# --- CARICAMENTO FILE ---
# ... (Codice invariato v17.9) ...
st.markdown("---"); st.markdown("#### 1. Carica la Baseline di Riferimento")
uploader_key = f"file_uploader_{st.session_state.widget_key_counter}"
uploaded_file = st.file_uploader("Seleziona il file .XML...", type=["xml"], label_visibility="collapsed", key=uploader_key)
if st.session_state.get('file_processed_success', False) and 'uploaded_file_state' in st.session_state : st.success('File XML analizzato con successo!')
if uploaded_file is not None and uploaded_file != st.session_state.get('uploaded_file_state'):
    st.session_state['uploaded_file_state'] = uploaded_file; st.session_state.file_processed_success = False
elif 'uploaded_file_state' not in st.session_state: uploaded_file = None

# --- FUNZIONI HELPER ---
# ... (get_minutes_per_day, format_duration_from_xml, get_tasks_to_distribute_for_sil, get_relevant_summary_name invariate) ...
# ... (classify_resource, extract_timephased_work invariate da v19.8) ...
@st.cache_data
def get_minutes_per_day(_tree, _ns):
    minutes_per_day = 480
    try:
        default_calendar = _tree.find(".//msp:Calendar[msp:UID='1']", namespaces=_ns)
        if default_calendar is not None:
            working_day = default_calendar.find(".//msp:WeekDay[msp:DayType='1']", namespaces=_ns)
            if working_day is not None:
                working_minutes = 0
                for working_time in working_day.findall(".//msp:WorkingTime", namespaces=_ns):
                    from_time_str = working_time.findtext('msp:FromTime', namespaces=_ns); to_time_str = working_time.findtext('msp:ToTime', namespaces=_ns)
                    if from_time_str and to_time_str:
                        try:
                            from_time = datetime.strptime(from_time_str, '%H:%M:%S').time(); to_time = datetime.strptime(to_time_str, '%H:%M:%S').time()
                            dummy_date = date(1, 1, 1); delta = datetime.combine(dummy_date, to_time) - datetime.combine(dummy_date, from_time)
                            working_minutes += delta.total_seconds() / 60
                        except ValueError: pass
                if working_minutes > 0: minutes_per_day = working_minutes
    except Exception: pass
    return minutes_per_day

def format_duration_from_xml(duration_str):
    minutes_per_day = st.session_state.get('minutes_per_day', 480)
    if not duration_str or minutes_per_day <= 0: return "0g"
    try:
        if duration_str.startswith('T'): duration_str = 'P' + duration_str
        elif not duration_str.startswith('P'): return "N/D"
        duration = isodate.parse_duration(duration_str); total_hours = duration.total_seconds() / 3600
        if total_hours == 0: return "0g"
        work_days = total_hours / (minutes_per_day / 60.0); return f"{round(work_days)}g"
    except Exception: return "N/D"

@st.cache_data
def get_tasks_to_distribute_for_sil(_tasks_dataframe):
    tasks_df = _tasks_dataframe.copy()
    tasks_df['Start'] = pd.to_datetime(tasks_df['Start'], errors='coerce').dt.date
    tasks_df['Finish'] = pd.to_datetime(tasks_df['Finish'], errors='coerce').dt.date
    tasks_df['WBS'] = tasks_df['WBS'].astype(str)
    valid_tasks_df = tasks_df.dropna(subset=['Start', 'Finish', 'Cost', 'WBS'])
    valid_tasks_df = valid_tasks_df[valid_tasks_df['Cost'] > 0]
    tasks_to_distribute_list = []
    processed_indices = set()
    for index, task in valid_tasks_df.iterrows():
        if index in processed_indices: continue
        task_wbs = task['WBS']
        has_child_with_cost = False
        for child_index, potential_child in valid_tasks_df.loc[valid_tasks_df.index != index].iterrows():
             child_wbs = potential_child['WBS']
             if child_wbs.startswith(task_wbs + '.') and child_wbs.count('.') == task_wbs.count('.') + 1:
                 has_child_with_cost = True; break
        if not has_child_with_cost:
            tasks_to_distribute_list.append(task.to_dict())
            processed_indices.add(index)
            descendant_indices = valid_tasks_df[valid_tasks_df['WBS'].str.startswith(task_wbs + '.')].index
            processed_indices.update(descendant_indices)
    if not tasks_to_distribute_list:
        tasks_to_distribute_df = pd.DataFrame()
    else:
        tasks_to_distribute_df = pd.DataFrame(tasks_to_distribute_list)
    st.session_state['debug_task_count'] = len(tasks_to_distribute_df)
    st.session_state['debug_total_cost'] = tasks_to_distribute_df['Cost'].sum() if not tasks_to_distribute_df.empty else 0
    return tasks_to_distribute_df

def get_relevant_summary_name(wbs_list, wbs_map):
    if not wbs_list: return "N/D"
    unique_wbs_list = sorted(list(set(wbs_list)))
    if len(unique_wbs_list) == 1:
        leaf_wbs = unique_wbs_list[0]
        if '.' in leaf_wbs:
            parent_wbs = leaf_wbs.rsplit('.', 1)[0]
            parent_name = wbs_map.get(parent_wbs)
            if parent_name: return parent_name
        return wbs_map.get(leaf_wbs, "Attività Sconosciuta")
    direct_parents = set()
    for wbs in unique_wbs_list:
        if '.' in wbs: direct_parents.add(wbs.rsplit('.', 1)[0])
        else: direct_parents.add(None)
    if len(direct_parents) == 1:
        parent_wbs = list(direct_parents)[0]
        if parent_wbs:
            parent_name = wbs_map.get(parent_wbs)
            if parent_name: return parent_name
    try:
        paths = [wbs.replace('.', '/') for wbs in unique_wbs_list]
        common_path_prefix = os.path.commonprefix(paths)
        if common_path_prefix.endswith('/'): common_path_prefix = common_path_prefix[:-1]
        common_wbs = common_path_prefix.replace('/', '.')
        if not common_wbs:
             root_task_name = wbs_map.get('1'); return root_task_name if root_task_name else "Riepilogo Progetto"
        parent_name = wbs_map.get(common_wbs)
        if parent_name: return parent_name
        else:
            parent_of_common = common_wbs.rsplit('.', 1)[0] if '.' in common_wbs else None
            if parent_of_common:
                grandparent_name = wbs_map.get(parent_of_common)
                if grandparent_name: return grandparent_name
            return f"Riepilogo: {common_wbs}"
    except Exception: return "Attività Multiple"

LABOR_KEYWORDS = [
    'operaio', 'ope ', 'addetto', 'squadra', 'assistente', 'tecnico', 'capo',
    'resp', 'ingegnere', 'geometra', 'sorvegliante', 'pilota', 'gruista',
    'autista', 'guardia', 'topografo', 'manovale', 'specializ', 'qualific',
    'comune', 'direttore', 'coordinatore', 'carpentiere', 'ferraiolo', 'mo'
]
EQUIPMENT_KEYWORDS = [
    'escavatore', 'pala', 'gru', 'terna', 'autocarro', 'camion', 'furgone',
    'mezzo', 'macchina', 'attrezz', 'pompa', 'generatore', 'compressore',
    'piattaforma', 'rullo', 'vibro', 'dumper', 'sonda', 'martello', 'tbm',
    'fresa', 'veicolo', 'auto', 'locomotore', 'carro', 'sollevatore', 'muletto',
    'mac', 'autogru', 'treno', 'posizionat', 'spritz', 'manitou', 'grader'
]

def classify_resource(resource_name):
    if not resource_name: return 'Altro'
    name_lower = resource_name.lower().strip()
    if any(keyword in name_lower for keyword in EQUIPMENT_KEYWORDS): return 'Mezzi'
    if any(keyword in name_lower for keyword in LABOR_KEYWORDS): return 'Manodopera'
    return 'Altro'

@st.cache_data
def extract_timephased_work(_xml_tree, _namespaces, _resource_map):
    daily_work_data = []
    assignments = _xml_tree.findall('.//msp:Assignment', namespaces=_namespaces)
    for ass in assignments:
        resource_uid_node = ass.find('msp:ResourceUID', namespaces=_namespaces)
        if resource_uid_node is None or resource_uid_node.text is None: continue
        resource_uid = resource_uid_node.text
        resource_name = _resource_map.get(resource_uid, '')
        resource_type = classify_resource(resource_name)
        timephased_nodes = ass.findall('./msp:TimephasedData[msp:Type="1"]', namespaces=_namespaces)
        for node in timephased_nodes:
            start_date_str = node.findtext('msp:Start', namespaces=_namespaces)
            value_str = node.findtext('msp:Value', namespaces=_namespaces)
            if start_date_str and value_str:
                try:
                    current_date = datetime.fromisoformat(start_date_str).date()
                    work_minutes = 0
                    if 'PT' in value_str:
                         duration_obj = isodate.parse_duration(value_str)
                         work_minutes = duration_obj.total_seconds() / 60.0
                    else: work_minutes = float(value_str)
                    if work_minutes > 0:
                        daily_work_data.append({'Date': current_date, 'ResourceUID': resource_uid, 'ResourceType': resource_type, 'WorkMinutes': work_minutes})
                except Exception: pass
    if not daily_work_data: return pd.DataFrame(columns=['Date', 'ResourceUID', 'ResourceType', 'WorkMinutes'])
    daily_df = pd.DataFrame(daily_work_data); daily_df['Date'] = pd.to_datetime(daily_df['Date'])
    return daily_df

# --- INIZIO ANALISI ---
current_file_to_process = st.session_state.get('uploaded_file_state')
if current_file_to_process is not None:
    if not st.session_state.get('file_processed_success', False) or current_file_to_process != st.session_state.get('last_processed_file'):
        with st.spinner('Caricamento e analisi file XML...'):
             try:
                # ... (Parsing XML, estrazione dati task, popolamento wbs_name_map invariato) ...
                current_file_to_process.seek(0); file_content_bytes = current_file_to_process.read()
                parser = etree.XMLParser(recover=True); tree = etree.fromstring(file_content_bytes, parser=parser)
                ns = {'msp': 'http://schemas.microsoft.com/project'}
                minutes_per_day = get_minutes_per_day(tree, ns); st.session_state['minutes_per_day'] = minutes_per_day
                project_name = "N/D"; formatted_cost = "€ 0,00"; project_start_date = None; project_finish_date = None
                task_uid_1 = tree.find(".//msp:Task[msp:UID='1']", namespaces=ns)
                if task_uid_1 is not None:
                    project_name = task_uid_1.findtext('msp:Name', namespaces=ns) or "N/D"; total_cost_str = task_uid_1.findtext('msp:Cost', namespaces=ns) or "0"; total_cost_euros = float(total_cost_str) / 100.0
                    formatted_cost = f"€ {total_cost_euros:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    start_str = task_uid_1.findtext('msp:Start', namespaces=ns); finish_str = task_uid_1.findtext('msp:Finish', namespaces=ns)
                    if start_str: project_start_date = datetime.fromisoformat(start_str).date()
                    if finish_str: project_finish_date = datetime.fromisoformat(finish_str).date()
                st.session_state['project_total_cost_from_summary'] = formatted_cost
                if not project_start_date: project_start_date = date.today()
                if not project_finish_date: project_finish_date = project_start_date + timedelta(days=365)
                if project_start_date > project_finish_date: project_finish_date = project_start_date + timedelta(days=1)
                st.session_state['project_name'] = project_name; st.session_state['formatted_cost'] = formatted_cost
                st.session_state['project_start_date'] = project_start_date; st.session_state['project_finish_date'] = project_finish_date
                potential_milestones = {}; all_tasks = tree.findall('.//msp:Task', namespaces=ns)
                tup_tuf_pattern = re.compile(r'(?i)(TUP|TUF)\s*\d*'); all_tasks_data_list = []
                wbs_name_map = {}
                for task in all_tasks:
                    uid = task.findtext('msp:UID', namespaces=ns); name = task.findtext('msp:Name', namespaces=ns) or "";
                    start_str = task.findtext('msp:Start', namespaces=ns); finish_str = task.findtext('msp:Finish', namespaces=ns);
                    start_date = datetime.fromisoformat(start_str).date() if start_str else None; finish_date = datetime.fromisoformat(finish_str).date() if finish_str else None
                    duration_str = task.findtext('msp:Duration', namespaces=ns);
                    cost_str = task.findtext('msp:Cost', namespaces=ns) or "0"
                    cost_euros = float(cost_str) / 100.0
                    duration_formatted = format_duration_from_xml(duration_str)
                    is_milestone_text = (task.findtext('msp:Milestone', namespaces=ns) or '0').lower(); is_milestone = is_milestone_text == '1' or is_milestone_text == 'true'
                    wbs = task.findtext('msp:WBS', namespaces=ns) or ""
                    total_slack_minutes_str = task.findtext('msp:TotalSlack', namespaces=ns) or "0"
                    is_summary_str = task.findtext('msp:Summary', namespaces=ns) or '0'
                    is_summary = is_summary_str == '1'
                    total_slack_days = 0
                    if total_slack_minutes_str:
                        try:
                            slack_minutes = float(total_slack_minutes_str)
                            mpd = st.session_state.get('minutes_per_day', 480)
                            if mpd > 0: total_slack_days = math.ceil(slack_minutes / mpd)
                        except ValueError: total_slack_days = 0
                    if wbs and name: wbs_name_map[wbs] = name
                    if uid != '0':
                        all_tasks_data_list.append({"UID": uid, "Name": name, "Start": start_date, "Finish": finish_date, "Duration": duration_formatted,
                                                    "Cost": cost_euros, "Milestone": is_milestone, "Summary": is_summary,
                                                    "WBS": wbs, "TotalSlackDays": total_slack_days})
                    match = tup_tuf_pattern.search(name)
                    if match:
                        tup_tuf_key = match.group(0).upper().strip(); duration_str_tup = task.findtext('msp:Duration', namespaces=ns)
                        try:
                            _ds = duration_str_tup
                            if _ds and _ds.startswith('T'): _ds = 'P' + _ds
                            duration_obj = isodate.parse_duration(_ds) if _ds and _ds.startswith('P') else timedelta(); duration_seconds = duration_obj.total_seconds()
                        except Exception: duration_seconds = 0
                        is_pure_milestone_duration = (duration_seconds == 0)
                        start_date_formatted = start_date.strftime("%d/%m/%Y") if start_date else "N/D"
                        finish_date_formatted = finish_date.strftime("%d/%m/%Y") if finish_date else "N/D"
                        current_task_data = {"Nome Completo": name, "Data Inizio": start_date_formatted, "Data Fine": finish_date_formatted, "Durata": duration_formatted, "DurataSecondi": duration_seconds, "DataInizioObj": start_date}
                        existing_duration_seconds = potential_milestones.get(tup_tuf_key, {}).get("DurataSecondi", -1)
                        if tup_tuf_key not in potential_milestones: potential_milestones[tup_tuf_key] = current_task_data
                        elif not is_pure_milestone_duration:
                            if existing_duration_seconds == 0: potential_milestones[tup_tuf_key] = current_task_data
                            elif duration_seconds > existing_duration_seconds: potential_milestones[tup_tuf_key] = current_task_data
                st.session_state['wbs_name_map'] = wbs_name_map
                final_milestones_data = []
                for key in potential_milestones:
                    data = potential_milestones[key]
                    final_milestones_data.append({"Nome Completo": data.get("Nome Completo", ""), "Data Inizio": data.get("Data Inizio", "N/D"),
                                                "Data Fine": data.get("Data Fine", "N/D"), "Durata": data.get("Durata", "N/D"),
                                                "DataInizioObj": data.get("DataInizioObj")})
                if final_milestones_data:
                    df_milestones = pd.DataFrame(final_milestones_data)
                    min_date_for_sort = date.min
                    df_milestones['DataInizioObj'] = pd.to_datetime(df_milestones['DataInizioObj'], errors='coerce').dt.date
                    df_milestones['DataInizioObj'] = df_milestones['DataInizioObj'].fillna(min_date_for_sort)
                    df_milestones = df_milestones.sort_values(by="DataInizioObj").reset_index(drop=True)
                    st.session_state['df_milestones_display'] = df_milestones[['Nome Completo', 'Durata', 'Data Inizio', 'Data Fine']]
                else: st.session_state['df_milestones_display'] = None
                st.session_state['all_tasks_data'] = pd.DataFrame(all_tasks_data_list)
                resources = tree.findall('.//msp:Resource', namespaces=ns)
                resource_map = {res.findtext('msp:UID', namespaces=ns): res.findtext('msp:Name', namespaces=ns) or f"Risorsa UID {res.findtext('msp:UID', namespaces=ns)}" for res in resources if res.findtext('msp:UID', namespaces=ns)}
                st.session_state['resource_map'] = resource_map
                timephased_work_data = extract_timephased_work(tree, ns, resource_map)
                st.session_state['timephased_work_data'] = timephased_work_data
                resource_class_list = []
                for uid, name in resource_map.items():
                    res_type = classify_resource(name)
                    resource_class_list.append({'UID': uid, 'Nome': name, 'Tipo Classificato': res_type})
                st.session_state['resource_classification_debug'] = pd.DataFrame(resource_class_list)
                current_file_to_process.seek(0); debug_content_bytes = current_file_to_process.read(2000);
                try: st.session_state['debug_raw_text'] = '\n'.join(debug_content_bytes.decode('utf-8', errors='ignore').splitlines()[:50])
                except Exception as decode_err: st.session_state['debug_raw_text'] = f"Errore decodifica debug: {decode_err}"
                st.session_state['last_processed_file'] = current_file_to_process
                st.session_state.file_processed_success = True
             except Exception as e:
                print(f"Errore Analisi: {e}"); print(traceback.format_exc())
                st.error(f"Errore Analisi durante elaborazione iniziale: {e}");
                st.error(f"Traceback: {traceback.format_exc()}");
                st.error("Verifica file XML.");
                st.session_state.file_processed_success = False;
                st.session_state['last_processed_file'] = None

    # --- VISUALIZZAZIONE DATI E ANALISI AVANZATA ---
    if st.session_state.get('file_processed_success', False):

        # --- Sezione 2 (Invariata) ---
        st.markdown("---"); st.markdown("#### 2. Analisi Preliminare"); st.markdown("##### 📄 Informazioni Generali dell'Appalto")
        project_name = st.session_state.get('project_name', "N/D"); formatted_cost = st.session_state.get('formatted_cost', "N/D")
        col1_disp, col2_disp = st.columns(2);
        with col1_disp: st.markdown(f"**Nome:** {project_name}")
        with col2_disp: st.markdown(f"**Importo Totale Lavori:** {formatted_cost}")
        st.markdown("##### 🗓️ Termini Utili Contrattuali (TUP/TUF)")
        df_display = st.session_state.get('df_milestones_display')
        if df_display is not None and not df_display.empty:
            st.dataframe(df_display, use_container_width=True, hide_index=True)
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer: df_display.to_excel(writer, index=False, sheet_name='TerminiUtili')
            excel_data = output.getvalue(); st.download_button(label="Scarica TUP/TUF (Excel)", data=excel_data, file_name="termini_utili_TUP_TUF.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="download_tup")
        else: st.warning("Nessun Termine Utile (TUP o TUF) trovato nel file.")


        # --- Sezione 3: Selezione Periodo e Analisi ---
        st.markdown("---"); st.markdown("#### 3. Analisi Avanzata")
        default_start = st.session_state.get('project_start_date', date.today()); default_finish = st.session_state.get('project_finish_date', date.today() + timedelta(days=365))
        if not default_start: default_start = date.today()
        if not default_finish: default_finish = default_start + timedelta(days=365)
        if default_start > default_finish: default_finish = default_start + timedelta(days=1)
        st.markdown("##### 📅 Seleziona Periodo di Riferimento"); st.caption(f"Default: {default_start.strftime('%d/%m/%Y')} - {default_finish.strftime('%d/%m/%Y')}.")
        col_date1, col_date2 = st.columns(2)
        with col_date1: selected_start_date = st.date_input("Data Inizio", value=default_start, min_value=default_start, max_value=default_finish + timedelta(days=5*365), format="DD/MM/YYYY", key="start_date_selector")
        with col_date2:
            min_end_date = selected_start_date; actual_default_finish = max(default_finish, min_end_date)
            reasonable_max_date = actual_default_finish + timedelta(days=10*365)
            selected_finish_date = st.date_input("Data Fine", value=actual_default_finish, min_value=min_end_date, max_value=reasonable_max_date, format="DD/MM/YYYY", key="finish_date_selector")
        st.markdown("##### 📦 Seleziona Aggregazione Dati")
        aggregation_level = st.radio("Scegli il livello di dettaglio per l'analisi:", ('Mensile', 'Giornaliera'), key="aggregation_selector", horizontal=True, help="Scegli 'Giornaliera' per visualizzare i dettagli giornalieri.")


        # --- Analisi Dettagliate ---
        st.markdown("---"); st.markdown("##### 📊 Analisi Dettagliate")

        # --- Analisi Curva S (Codice invariato da v18.3) ---
        if st.button("📈 Avvia Analisi Curva S", key="analyze_scurve"):
            all_tasks_dataframe = st.session_state.get('all_tasks_data'); wbs_name_map = st.session_state.get('wbs_name_map', {})
            if all_tasks_dataframe is None or all_tasks_dataframe.empty: st.error("Errore: Dati attività non trovati.")
            elif not wbs_name_map: st.error("Errore: Mappa WBS->Nome non trovata.")
            else:
                try:
                    st.markdown(f"###### Analisi Curva S")
                    tasks_to_distribute = get_tasks_to_distribute_for_sil(all_tasks_dataframe.copy())
                    if tasks_to_distribute.empty: st.error("Errore: Nessun costo valido trovato per la distribuzione.")
                    else:
                        daily_cost_data = []; total_tasks = len(tasks_to_distribute); status_text = st.empty(); prog_bar = st.progress(0, text="Avvio calcolo distribuzione costi...")
                        for i, (_, task) in enumerate(tasks_to_distribute.iterrows()):
                            start_date_task = task['Start']; finish_date_task = task['Finish']; total_cost_task = task['Cost']; task_wbs = task['WBS']
                            duration_days = (finish_date_task - start_date_task).days
                            if duration_days < 0: continue
                            number_of_days_in_period = duration_days + 1
                            if number_of_days_in_period <= 0: value_per_day = total_cost_task; number_of_days_in_period = 1
                            else: value_per_day = total_cost_task / number_of_days_in_period
                            for d in range(number_of_days_in_period):
                                current_date = start_date_task + timedelta(days=d)
                                daily_cost_data.append({'Date': current_date, 'Value': value_per_day, 'WBS': task_wbs})
                            percentage = (i + 1) / total_tasks
                            prog_bar.progress(percentage, text=f"Calcolo distribuzione costi: {percentage:.0%}")
                        prog_bar.empty(); status_text.empty()
                        if not daily_cost_data: st.error("Errore: Nessun dato di costo generato.")
                        else:
                            detailed_daily_cost_df = pd.DataFrame(daily_cost_data); detailed_daily_cost_df['Date'] = pd.to_datetime(detailed_daily_cost_df['Date'])
                            aggregated_daily_raw = detailed_daily_cost_df.groupby('Date').agg(Value=('Value', 'sum'), WBS_List=('WBS', lambda x: list(set(x)))).reset_index()
                            selected_start_dt = datetime.combine(selected_start_date, datetime.min.time()); selected_finish_dt = datetime.combine(selected_finish_date, datetime.max.time())
                            mask_cost = (aggregated_daily_raw['Date'] >= selected_start_dt) & (aggregated_daily_raw['Date'] <= selected_finish_dt)
                            filtered_cost = aggregated_daily_raw.loc[mask_cost]
                            if not filtered_cost.empty:
                                aggregated_data = pd.DataFrame(); display_columns = []; plot_custom_data = None; col_summary_name = "Riepilogo WBS"; date_format_display = ""; date_format_excel = ""; excel_filename = ""
                                if aggregation_level == 'Mensile':
                                    aggregated_values = filtered_cost.set_index('Date')['Value'].resample('ME').sum().reset_index()
                                    aggregated_values = aggregated_values.sort_values(by='Date') # <<< Ordina
                                    aggregated_data = aggregated_values
                                    date_format_display = '%b-%y'; date_format_excel = '%b-%y'
                                    aggregated_data['Periodo'] = aggregated_data['Date'].dt.strftime(date_format_display).str.capitalize()
                                    axis_title = "Mese"; col_name = "Costo Mensile (€)"; display_columns = ['Periodo', col_name, 'Costo Cumulato (€)']; excel_filename = "Dati_SIL_Mensili.xlsx"
                                else: # Giornaliera
                                    aggregated_daily = filtered_cost.copy(); aggregated_daily[col_summary_name] = aggregated_daily['WBS_List'].apply(lambda l: get_relevant_summary_name(l, wbs_name_map))
                                    aggregated_data = aggregated_daily; date_format_display = '%d/%m/%Y'; date_format_excel = '%d/%m/%Y'; aggregated_data['Periodo'] = aggregated_data['Date'].dt.strftime(date_format_display)
                                    axis_title = "Giorno"; col_name = "Costo Giornaliero (€)"; display_columns = ['Periodo', col_name, 'Costo Cumulato (€)', col_summary_name]; plot_custom_data = aggregated_data[col_summary_name]; excel_filename = "Dati_SIL_Giornalieri.xlsx"
                                aggregated_data['Costo Cumulato (€)'] = aggregated_data['Value'].cumsum()
                                st.markdown(f"###### Tabella Dati SIL Aggregati ({aggregation_level})"); df_display_sil = aggregated_data.copy(); df_display_sil.rename(columns={'Value': col_name}, inplace=True)
                                df_display_sil[col_name] = df_display_sil[col_name].apply(lambda x: f"€ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")); df_display_sil['Costo Cumulato (€)'] = df_display_sil['Costo Cumulato (€)'].apply(lambda x: f"€ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
                                st.dataframe(df_display_sil[display_columns], use_container_width=True, hide_index=True)
                                st.markdown(f"###### Grafico Curva S ({aggregation_level})"); fig_sil = go.Figure()
                                hovertemplate_bar = f'<b>{axis_title}</b>: %{{x}}<br><b>Costo {aggregation_level}</b>: %{{y:,.2f}}€<extra></extra>'; hovertemplate_scatter = f'<b>{axis_title}</b>: %{{x}}<br><b>Costo Cumulato</b>: %{{y:,.2f}}€<extra></extra>'
                                if aggregation_level == 'Giornaliera': hovertemplate_bar = f'<b>{axis_title}</b>: %{{x}}<br><b>Costo {col_name}</b>: %{{y:,.2f}}€<br><b>{col_summary_name}</b>: %{{customdata}}<extra></extra>'; hovertemplate_scatter = f'<b>{axis_title}</b>: %{{x}}<br><b>Costo Cumulato</b>: %{{y:,.2f}}€<br><b>{col_summary_name}</b>: %{{customdata}}<extra></extra>'
                                fig_sil.add_trace(go.Bar(x=aggregated_data['Periodo'], y=aggregated_data['Value'], name=f'Costo {aggregation_level}', customdata=plot_custom_data, hovertemplate=hovertemplate_bar, marker_color='royalblue'))
                                fig_sil.add_trace(go.Scatter(x=aggregated_data['Periodo'], y=aggregated_data['Costo Cumulato (€)'], name=f'Costo Cumulato', mode='lines+markers', yaxis='y2', customdata=plot_custom_data, hovertemplate=hovertemplate_scatter, line_color='crimson', marker_color='crimson'))
                                fig_sil.update_layout(title=f'Curva S - Costo {aggregation_level.replace("a", "o")} e Cumulato', xaxis_title=axis_title, yaxis=dict(title=f"Costo {aggregation_level.replace('a', 'o')} (€)"), yaxis2=dict(title="Costo Cumulato (€)", overlaying="y", side="right"), legend=dict(yanchor="top", y=0.99, xanchor="left", x=0.01), hovermode="x unified", template="plotly")
                                st.plotly_chart(fig_sil, use_container_width=True)
                                output_sil = BytesIO(); df_export = aggregated_data.copy(); cols_to_select_excel = []; rename_map_excel = {}; excel_sheet_name = 'Tabella'
                                if aggregation_level == 'Mensile': cols_to_select_excel = ['Date', 'Value', 'Costo Cumulato (€)']; rename_map_excel = {'Date': 'Mese', 'Value': 'Costo Mensile (€)'}; df_export['Date'] = df_export['Date'].dt.strftime(date_format_excel).str.capitalize()
                                else: cols_to_select_excel = ['Date', 'Value', 'Costo Cumulato (€)', col_summary_name]; rename_map_excel = {'Date': 'Giorno', 'Value': 'Costo Giornaliero (€)', col_summary_name: 'Riepilogo WBS'}; df_export['Date'] = df_export['Date'].dt.strftime(date_format_excel)
                                df_to_write = df_export[cols_to_select_excel]; df_to_write = df_to_write.rename(columns=rename_map_excel)
                                with pd.ExcelWriter(output_sil, engine='openpyxl') as writer:
                                    df_to_write.to_excel(writer, index=False, sheet_name=excel_sheet_name); worksheet_table = writer.sheets[excel_sheet_name]
                                    for idx, col in enumerate(df_to_write):
                                        try: series = df_to_write[col]; max_len = max((series.astype(str).map(len).max(), len(str(series.name)))) + 3; worksheet_table.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = max_len
                                        except Exception as cw_err: print(f"Err col {col}: {cw_err}")
                                    if _kaleido_installed:
                                        try: img_bytes = pio.to_image(fig_sil, format="png", width=900, height=500, scale=1.5); img = Image(BytesIO(img_bytes)); worksheet_chart = writer.book.create_sheet(title='Grafico'); worksheet_chart.add_image(img, 'A1')
                                        except Exception as img_err: st.warning(f"Impossibile esportare il grafico in Excel (errore Kaleido/Plotly): {img_err}")
                                    else: st.warning("Kaleido mancante.")
                                excel_data_sil = output_sil.getvalue()
                                st.download_button(label=f"Scarica SIL ({aggregation_level})", data=excel_data_sil, file_name=excel_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="download_sil")
                                st.markdown("---"); st.markdown(f"##### Diagnostica Dati Calcolati"); debug_task_count = st.session_state.get('debug_task_count', 0); st.write(f"**N. attività usate:** {debug_task_count}"); debug_total = st.session_state.get('debug_total_cost', 0); formatted_debug_cost = f"€ {debug_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."); st.write(f"**Costo Totale Calcolato:** {formatted_debug_cost}"); project_total = st.session_state.get('project_total_cost_from_summary', 'N/D'); st.caption(f"Costo Totale Ufficiale: {project_total}"); st.caption("I totali dovrebbero corrispondere.")
                            else: st.warning(f"Nessun dato di costo trovato nel periodo selezionato.")
                except Exception as analysis_error: st.error(f"Errore Analisi Avanzata: {analysis_error}"); st.error(traceback.format_exc())

        # --- [MODIFICATO v19.12] Sezione Istogrammi Risorse ---
        st.markdown("---")
        st.markdown("###### 📊 Istogrammi Risorse (Unità Medie Giornaliere eq. 8h)")

        # --- [MODIFICATO v19.12] Rimossa opzione "Tutte" ---
        resource_type_options = ['Manodopera', 'Mezzi', 'Altro']
        selected_resource_type = st.selectbox(
            "Seleziona il tipo di risorsa da analizzare:",
            resource_type_options,
            key="resource_type_selector",
            help="Mostra le unità medie giornaliere equivalenti (Manodopera = totale, Mezzi/Altro = dettaglio)."
        )

        if st.button("📊 Avvia Analisi Istogrammi", key="analyze_histograms"):
            timephased_work_df = st.session_state.get('timephased_work_data')
            resource_map = st.session_state.get('resource_map', {})

            if timephased_work_df is None or timephased_work_df.empty:
                st.error("Errore: Dati 'Timephased Work' non trovati nel file XML.")
                st.warning("Assicurati che il progetto abbia risorse assegnate alle attività con del lavoro pianificato o effettivo.")
            elif not resource_map:
                st.error("Errore: Mappa Risorse non trovata.")
            else:
                try:
                    with st.spinner(f"Calcolo unità medie giornaliere ({selected_resource_type})..."):
                        work_df_filtered = timephased_work_df.copy()
                        work_df_filtered = work_df_filtered[work_df_filtered['ResourceType'] == selected_resource_type]

                        selected_start_dt = datetime.combine(selected_start_date, datetime.min.time())
                        selected_finish_dt = datetime.combine(selected_finish_date, datetime.max.time())
                        mask_work = (work_df_filtered['Date'] >= selected_start_dt) & (work_df_filtered['Date'] <= selected_finish_dt)
                        filtered_work = work_df_filtered.loc[mask_work].copy()

                        if filtered_work.empty:
                             st.warning(f"Nessun dato di lavoro trovato per '{selected_resource_type}' nel periodo selezionato.")
                        else:
                            filtered_work['WorkHours'] = filtered_work['WorkMinutes'] / 60.0
                            filtered_work['ResourceName'] = filtered_work['ResourceUID'].map(resource_map).fillna('Sconosciuto')

                            date_format_display_hist = '%b-%y' if aggregation_level == 'Mensile' else '%d/%m/%Y'
                            date_format_excel_hist = '%b-%y' if aggregation_level == 'Mensile' else '%d/%m/%Y'
                            axis_title_hist = "Mese" if aggregation_level == 'Mensile' else "Giorno"
                            col_name_hist = f"Unità Media Giorn."
                            excel_filename_hist = f"Istogramma_UnitaMediaGiorn_{selected_resource_type.replace(' ', '_')}_{aggregation_level}.xlsx"
                            df_pivot_export = None

                            # --- [MODIFICATO v19.12] Logica differenziata (Mezzi e Altro vs Manodopera) ---
                            if selected_resource_type in ['Mezzi', 'Altro']:
                                # Dettaglio per Risorsa (Mezzi o Altro)
                                aggregated_daily_detail = filtered_work.groupby(['Date', 'ResourceName'])['WorkHours'].sum().reset_index()

                                if aggregation_level == 'Mensile':
                                    aggregated_hist_raw = aggregated_daily_detail.set_index('Date').groupby('ResourceName')['WorkHours'].resample('ME').sum().reset_index()
                                    aggregated_hist_raw['DaysInMonth'] = aggregated_hist_raw['Date'].dt.daysinmonth
                                    aggregated_hist = aggregated_hist_raw
                                    aggregated_hist['AvgDailyUnits'] = (aggregated_hist['WorkHours'] / 8.0) / aggregated_hist['DaysInMonth']
                                else: # Giornaliera
                                    aggregated_hist = aggregated_daily_detail
                                    aggregated_hist['AvgDailyUnits'] = aggregated_hist['WorkHours'] / 8.0

                                aggregated_hist = aggregated_hist.sort_values(by=['Date', 'ResourceName'])
                                aggregated_hist['Periodo'] = aggregated_hist['Date'].dt.strftime(date_format_display_hist).str.capitalize()
                                aggregated_hist['AvgDailyUnits_Rounded'] = aggregated_hist['AvgDailyUnits'].round().astype(int)

                                # --- VISUALIZZAZIONE MEZZI / ALTRO ---
                                st.markdown(f"###### Tabella Dettaglio Unità Medie Giorn. {selected_resource_type} ({aggregation_level})")
                                df_display_hist = aggregated_hist.copy()
                                df_display_hist.rename(columns={'AvgDailyUnits_Rounded': col_name_hist}, inplace=True)
                                
                                try:
                                    pivot_table = pd.pivot_table(df_display_hist, values=col_name_hist, index='Periodo', columns='ResourceName', aggfunc='first', fill_value=0)
                                    pivot_table = pivot_table.reindex(aggregated_hist['Periodo'].unique()) # Forza ordinamento cronologico
                                    st.dataframe(pivot_table, use_container_width=True)
                                except Exception as e_pivot:
                                    st.warning(f"Impossibile creare tabella pivot ({e_pivot}). Mostro tabella standard.")
                                    st.dataframe(df_display_hist[['Periodo', 'ResourceName', col_name_hist]].sort_values(by=['Date', 'ResourceName']), use_container_width=True, hide_index=True)

                                st.markdown(f"###### Grafico Istogramma Unità Medie Giorn. {selected_resource_type} ({aggregation_level})")
                                fig_hist = go.Figure()
                                colors = px.colors.qualitative.Plotly
                                resource_names = aggregated_hist['ResourceName'].unique()
                                for i, name in enumerate(resource_names):
                                    group = aggregated_hist[aggregated_hist['ResourceName'] == name]
                                    fig_hist.add_trace(go.Bar(
                                        x=group['Periodo'],
                                        y=group['AvgDailyUnits_Rounded'],
                                        name=name,
                                        marker_color=colors[i % len(colors)],
                                        hovertemplate=f'<b>{axis_title_hist}</b>: %{{x}}<br><b>Risorsa</b>: {name}<br><b>Unità Media Giorn.</b>: %{{y:,.0f}}<extra></extra>'
                                    ))
                                fig_hist.update_layout(
                                    title=f'Istogramma Unità Medie Giorn. {selected_resource_type} - {aggregation_level.replace("a","e")} per Risorsa',
                                    xaxis_title=axis_title_hist,
                                    yaxis=dict(title=f"Unità Medie Giorn. {aggregation_level.replace('a','e')} (eq. 8h)"),
                                    barmode='group',
                                    legend=dict(yanchor="top", y=0.99, xanchor="left", x=0.01),
                                    hovermode="x unified",
                                    template="plotly",
                                    yaxis_tickformat = ',.0f'
                                )
                                st.plotly_chart(fig_hist, use_container_width=True)

                                # --- EXPORT EXCEL MEZZI / ALTRO ---
                                output_hist = BytesIO()
                                df_export_hist = aggregated_hist.copy()
                                rename_map_excel_hist = {'Date': axis_title_hist, 'AvgDailyUnits_Rounded': col_name_hist, 'ResourceName': 'Risorsa'}
                                df_export_hist['Date'] = df_export_hist['Date'].dt.strftime(date_format_excel_hist).str.capitalize() if aggregation_level=='Mensile' else df_export_hist['Date'].dt.strftime(date_format_excel_hist)
                                df_to_write_hist = df_export_hist[['Date', 'ResourceName', 'AvgDailyUnits_Rounded']]
                                df_to_write_hist = df_to_write_hist.rename(columns=rename_map_excel_hist)

                                if aggregation_level == 'Mensile':
                                     try:
                                        df_pivot_export = pd.pivot_table(df_to_write_hist, values=col_name_hist, index=axis_title_hist, columns='Risorsa', fill_value=0)
                                        # Forza ordinamento corretto anche in Excel
                                        df_pivot_export = df_pivot_export.reindex(df_export_hist['Date'].unique())
                                     except Exception: df_pivot_export = None

                            else: # Manodopera (Totale)
                                aggregated_daily_total = filtered_work.groupby('Date')['WorkHours'].sum().reset_index()

                                if aggregation_level == 'Mensile':
                                    aggregated_hist_raw = aggregated_daily_total.set_index('Date')['WorkHours'].resample('ME').sum().reset_index()
                                    aggregated_hist_raw = aggregated_hist_raw.sort_values(by='Date') # <<< Ordinamento
                                    aggregated_hist_raw['DaysInMonth'] = aggregated_hist_raw['Date'].dt.daysinmonth
                                    aggregated_hist = aggregated_hist_raw
                                    aggregated_hist['AvgDailyUnits'] = (aggregated_hist['WorkHours'] / 8.0) / aggregated_hist['DaysInMonth']
                                else: # Giornaliera
                                    aggregated_hist = aggregated_daily_total
                                    aggregated_hist['AvgDailyUnits'] = aggregated_hist['WorkHours'] / 8.0

                                aggregated_hist['Periodo'] = aggregated_hist['Date'].dt.strftime(date_format_display_hist).str.capitalize()
                                aggregated_hist['AvgDailyUnits_Rounded'] = aggregated_hist['AvgDailyUnits'].round().astype(int)

                                # --- VISUALIZZAZIONE MANODOPERA/TUTTE ---
                                st.markdown(f"###### Tabella Totale Unità Medie Giorn. {selected_resource_type} ({aggregation_level})")
                                df_display_hist = aggregated_hist.copy()
                                df_display_hist.rename(columns={'AvgDailyUnits_Rounded': col_name_hist}, inplace=True)
                                st.dataframe(df_display_hist[['Periodo', col_name_hist]], use_container_width=True, hide_index=True)

                                st.markdown(f"###### Grafico Istogramma Totale Unità Medie Giorn. {selected_resource_type} ({aggregation_level})")
                                aggregated_hist_plot = aggregated_hist.copy()
                                fig_hist = go.Figure()
                                fig_hist.add_trace(go.Bar(
                                    x=aggregated_hist_plot['Periodo'],
                                    y=aggregated_hist_plot['AvgDailyUnits_Rounded'],
                                    name=f'Unità Media Giorn. {aggregation_level}',
                                    marker_color='mediumseagreen',
                                    hovertemplate=f'<b>{axis_title_hist}</b>: %{{x}}<br><b>Unità Media Giorn.</b>: %{{y:,.0f}}<extra></extra>'
                                ))
                                fig_hist.update_layout(
                                    title=f'Istogramma Totale Unità Medie Giorn. ({selected_resource_type}) - {aggregation_level.replace("a","e")}',
                                    xaxis_title=axis_title_hist,
                                    yaxis=dict(title=f"Unità Medie Giorn. {aggregation_level.replace('a','e')} (eq. 8h)"),
                                    hovermode="x unified",
                                    template="plotly",
                                    yaxis_tickformat = ',.0f'
                                )
                                st.plotly_chart(fig_hist, use_container_width=True)

                                # --- EXPORT EXCEL MANODOPERA/TUTTE ---
                                output_hist = BytesIO()
                                df_export_hist = aggregated_hist.copy()
                                rename_map_excel_hist = {'Date': axis_title_hist, 'AvgDailyUnits_Rounded': col_name_hist}
                                df_export_hist['Date'] = df_export_hist['Date'].dt.strftime(date_format_excel_hist).str.capitalize() if aggregation_level=='Mensile' else df_export_hist['Date'].dt.strftime(date_format_excel_hist)
                                df_to_write_hist = df_export_hist[['Date', 'AvgDailyUnits_Rounded']]
                                df_to_write_hist = df_to_write_hist.rename(columns=rename_map_excel_hist)

                            # --- Export Excel (Comune) ---
                            with pd.ExcelWriter(output_hist, engine='openpyxl') as writer:
                                if selected_resource_type in ['Mezzi', 'Altro'] and aggregation_level == 'Mensile' and df_pivot_export is not None:
                                    df_pivot_export.to_excel(writer, sheet_name='Tabella_Pivot')
                                    worksheet_pivot = writer.sheets['Tabella_Pivot']
                                    for idx, col in enumerate(df_pivot_export.columns):
                                        col_letter = openpyxl.utils.get_column_letter(idx + 2);
                                        try:
                                            col_str = str(col)
                                            max_len = max((df_pivot_export[col].astype(str).map(len).max(), len(col_str))) + 3
                                        except: max_len = len(str(col)) + 3
                                        worksheet_pivot.column_dimensions[col_letter].width = max_len
                                    try:
                                        idx_name_len = len(str(df_pivot_export.index.name)) if df_pivot_export.index.name else 0
                                        idx_val_len = df_pivot_export.index.astype(str).map(len).max()
                                        idx_len = max(idx_name_len, idx_val_len) + 3
                                    except: idx_len = 15
                                    worksheet_pivot.column_dimensions['A'].width = idx_len
                                else:
                                    df_to_write_hist.to_excel(writer, index=False, sheet_name='Tabella')
                                    worksheet_table_hist = writer.sheets['Tabella']
                                    for idx, col in enumerate(df_to_write_hist):
                                        try:
                                            series = df_to_write_hist[col]
                                            max_len = max((series.astype(str).map(len).max(), len(str(series.name)))) + 3
                                            worksheet_table_hist.column_dimensions[openpyxl.utils.get_column_letter(idx + 1)].width = max_len
                                        except Exception as col_width_err: print(f"Errore agg colonna hist {col}: {col_width_err}")

                                if _kaleido_installed:
                                    try:
                                        img_bytes_hist = pio.to_image(fig_hist, format="png", width=900, height=500, scale=1.5)
                                        img_hist = Image(BytesIO(img_bytes_hist))
                                        worksheet_chart_hist = writer.book.create_sheet(title='Grafico')
                                        worksheet_chart_hist.add_image(img_hist, 'A1')
                                    except Exception as img_err_hist: st.warning(f"Impossibile esportare grafico istogramma: {img_err_hist}")
                                else: st.warning("Kaleido mancante per export grafico istogramma.")

                            excel_data_hist = output_hist.getvalue()
                            st.download_button(label=f"Scarica Istogramma ({aggregation_level}, {selected_resource_type})", data=excel_data_hist, file_name=excel_filename_hist, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="download_hist")


                except Exception as analysis_error_hist:
                    st.error(f"Errore during l'analisi degli istogrammi: {analysis_error_hist}")
                    st.error(traceback.format_exc())
        
        # --- [NUOVO v20.1] Sezione Analisi Percorso Critico ---
        st.markdown("---")
        st.markdown("###### ⛓️ Analisi Percorso Critico")
        
        st.caption("""
        Il **Margine di Flessibilità Totale** (Total Slack) indica di quanto tempo un'attività può ritardare senza influenzare la data di fine totale del progetto.
        Un'attività è considerata **critica** se ha un margine di flessibilità pari o inferiore a 0 giorni.
        Impostando un valore (es. 5 giorni), puoi identificare anche le attività **quasi-critiche**.
        """)
        
        slack_threshold = st.number_input(
            "Mostra attività con Flessibilità Totale (giorni) minore o uguale a:",
            min_value=0, max_value=100, value=0, step=1,
            key="slack_threshold_selector",
            help="Default = 0 (percorso critico stretto). Aumenta per includere attività quasi-critiche."
        )

        if st.button("🔬 Avvia Analisi Criticità", key="analyze_critical_path"):
            all_tasks_df = st.session_state.get('all_tasks_data')
            
            if all_tasks_df is None or all_tasks_df.empty:
                st.error("Errore: Dati delle attività non trovati.")
            else:
                try:
                    with st.spinner(f"Calcolo attività critiche (Flessibilità <= {slack_threshold} giorni)..."):
                        
                        tasks_df_crit = all_tasks_df.copy()
                        # Assicura che le date siano nel formato corretto
                        tasks_df_crit['Start'] = pd.to_datetime(tasks_df_crit['Start'], errors='coerce').dt.date
                        tasks_df_crit['Finish'] = pd.to_datetime(tasks_df_crit['Finish'], errors='coerce').dt.date
                        
                        # Filtro 1: Non di riepilogo
                        tasks_df_crit = tasks_df_crit[tasks_df_crit['Summary'] == False]
                        
                        # Filtro 2: Flessibilità Totale
                        tasks_df_crit = tasks_df_crit[tasks_df_crit['TotalSlackDays'] <= slack_threshold]
                        
                        # Filtro 3: Sovrapposizione con periodo selezionato
                        mask_overlap = (tasks_df_crit['Start'] <= selected_finish_date) & (tasks_df_crit['Finish'] >= selected_start_date)
                        critical_tasks_in_period = tasks_df_crit[mask_overlap]

                    if critical_tasks_in_period.empty:
                        st.warning(f"Nessuna attività (non di riepilogo) trovata con Flessibilità Totale <= {slack_threshold} giorni nel periodo selezionato.")
                    else:
                        st.markdown(f"###### Attività Critiche e Quasi-Critiche nel Periodo (Flessibilità <= {slack_threshold} giorni)")
                        
                        df_display_crit = critical_tasks_in_period.copy()
                        df_display_crit['Start_dt'] = pd.to_datetime(df_display_crit['Start']) # Colonna per ordinamento
                        df_display_crit['Start'] = df_display_crit['Start'].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else 'N/D')
                        df_display_crit['Finish'] = df_display_crit['Finish'].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else 'N/D')
                        
                        # --- [MODIFICATO v20.2] Ordine colonne Percorso Critico ---
                        cols_to_show = ['Name', 'Duration', 'Start', 'Finish', 'WBS', 'TotalSlackDays']
                        st.dataframe(df_display_crit.sort_values(by='Start_dt')[cols_to_show], use_container_width=True, hide_index=True)

                        # Bottone Download
                        output_crit = BytesIO()
                        with pd.ExcelWriter(output_crit, engine='openpyxl') as writer:
                            # Esporta con le colonne nell'ordine richiesto
                            df_display_crit.sort_values(by='Start_dt')[cols_to_show].to_excel(writer, index=False, sheet_name='Attivita_Critiche')
                        excel_data_crit = output_crit.getvalue()
                        # --- FINE MODIFICA ---

                        st.download_button(
                            label=f"Scarica Attività Critiche (Excel)",
                            data=excel_data_crit,
                            file_name=f"attivita_critiche_slack{slack_threshold}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_critical"
                        )
                        
                except Exception as analysis_error_crit:
                    st.error(f"Errore durante l'analisi del percorso critico: {analysis_error_crit}")
                    st.error(traceback.format_exc())
        # --- FINE NUOVA SEZIONE ---

        # --- [MODIFICATO v20.2] Debug spostati qui (indentazione corretta) ---
        st.markdown("---")
        with st.expander("🔍 Area Debug (Avanzato)", collapsed=True):
            st.markdown("##### Debug: Classificazione Risorse")
            df_res_class = st.session_state.get('resource_classification_debug')
            if df_res_class is not None and not df_res_class.empty:
                st.write("Elenco di tutte le risorse trovate e come sono state classificate (Logica: Mezzi prima di Manodopera):")
                st.dataframe(df_res_class, use_container_width=True, height=300, hide_index=True)
                counts = df_res_class['Tipo Classificato'].value_counts().reset_index()
                counts.columns = ['Tipo', 'Conteggio']
                st.write("Conteggio Totale per Tipo:")
                st.dataframe(counts, hide_index=True)
            else:
                st.warning("Nessuna risorsa trovata o mappa non generata.")

            st.markdown("---")
            st.markdown("##### Dati Grezzi per Debug (prime 50 righe del file)")
            debug_text = st.session_state.get('debug_raw_text')
            if debug_text:
                st.code(debug_text, language='xml')
            else:
                st.info("Dati grezzi non disponibili.")
        # --- FINE MODIFICA ---
